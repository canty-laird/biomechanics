#!/usr/bin/env python3
# biomechanics_ddmmyy.n.py
#
# 'Age' is not coerced to numeric (it stays as '8wks', '18wks', '52wks').

import os
import re
import numpy as np
import pandas as pd
from scipy.signal import savgol_filter

# imports for plotting + Excel embedding
import matplotlib
matplotlib.use("Agg")  # Non-interactive backend for safe batch plotting
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------
# Smoothing configuration
# ---------------------------------------------------------------------

PRECON_SAVGOL_WINDOW = 101
PRECON_SAVGOL_POLY   = 3
HOLD_SAVGOL_WINDOW   = 101
HOLD_SAVGOL_POLY     = 3
FAILURE_SAVGOL_WINDOW = 51
FAILURE_SAVGOL_POLY   = 3
USE_SAVGOL_FAILURE    = True

# ---------------------------------------------------------------------
# File utilities
# ---------------------------------------------------------------------

def find_data_files(root_dir):
    data_files = []
    for dirpath, _, filenames in os.walk(root_dir):
        for fname in filenames:
            if fname.endswith("Data.csv"):
                data_files.append(os.path.join(dirpath, fname))
    return sorted(data_files)

def parse_sample_from_filename(path):
    fname = os.path.basename(path)
    m = re.search(r"(\d{6}).*?Sample\s*([A-Z])\s*([0-9]+)", fname, flags=re.IGNORECASE)
    if not m:
        return None, None, None
    return m.group(1), m.group(2).upper(), m.group(3)

# ---------------------------------------------------------------------
# helper function to generate Force–Time plots
# ---------------------------------------------------------------------

def save_force_time_plot(df_full, outfile_png):
    """
    Save a small (approx 400x300 px) Force vs Time plot as PNG.
    Plots the *entire* time series (preconditioning + hold + failure).
    """
    if df_full is None or df_full.empty:
        return None

    try:
        times = df_full["Time_S"].to_numpy(float)
        force = df_full["Force_N"].to_numpy(float)
    except Exception:
        return None

    if len(times) < 5:
        return None

    plt.figure(figsize=(4, 3), dpi=120)
    plt.plot(times, force, linewidth=1)
    plt.xlabel("Time (s)")
    plt.ylabel("Force (N)")
    plt.title("Force vs Time (entire test)")
    plt.tight_layout()
    plt.savefig(outfile_png, dpi=120)
    plt.close()
    return outfile_png

# ---------------------------------------------------------------------
# helper function: Force–Extension plot data from which modulus is derived marked
# ---------------------------------------------------------------------

def save_force_extension_plot(df_fail, outfile_png, disp_mod=None, force_mod=None):
    """
    Plot Force vs Extension for the failure segment and optionally
    mark the modulus location with a red X.
    """
    if df_fail is None or df_fail.empty:
        return None

    try:
        disp = df_fail["Displacement_mm"].to_numpy(float)
        force = df_fail["Force_N"].to_numpy(float)
    except Exception:
        return None

    if len(disp) < 5:
        return None

    plt.figure(figsize=(4, 3), dpi=120)
    plt.plot(disp, force, linewidth=1)

    # Mark modulus point
    if disp_mod is not None and force_mod is not None:
        plt.scatter([disp_mod], [force_mod], s=50, marker="x", color="red")

    plt.xlabel("Extension (mm)")
    plt.ylabel("Force (N)")
    plt.title("Force vs Extension (failure segment)")
    plt.tight_layout()
    plt.savefig(outfile_png, dpi=120)
    plt.close()

    return outfile_png


# ---------------------------------------------------------------------
# Metadata
# ---------------------------------------------------------------------

def load_metadata(meta_path):
    if not os.path.exists(meta_path):
        raise FileNotFoundError(f"Metadata file '{meta_path}' not found.")

    df = pd.read_csv(meta_path)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    rename_map = {}
    for col in df.columns:
        cl = col.lower()
        if cl == "date_id":
            rename_map[col] = "Date_ID"
        elif cl == "sample_id":
            rename_map[col] = "Sample_ID"
        elif cl == "replicate":
            rename_map[col] = "Replicate"
        elif cl == "average_diameter (um)":
            rename_map[col] = "Average_diameter (um)"
        elif cl == "c.s.a (um squared)":
            rename_map[col] = "C.s.a (um squared)"
        elif cl == "c.s.a (mm squared)":
            rename_map[col] = "C.s.a (mm squared)"

    df = df.rename(columns=rename_map)

    # NOTE: 'Age' is intentionally NOT coerced to numeric, since it is categorical
    num_cols = ["Average_diameter (um)", "C.s.a (um squared)", "C.s.a (mm squared)"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    return df

# ---------------------------------------------------------------------
# Savitzky–Golay smoothing with safety
# ---------------------------------------------------------------------

def apply_savgol_safe(y, window, poly):
    y = np.asarray(y, dtype=float)
    n = len(y)
    if n < 5:
        return y.copy()
    w = min(window, n)
    if w % 2 == 0:
        w -= 1
    if w < 5:
        return y.copy()
    p = min(poly, w - 1)
    return savgol_filter(y, window_length=w, polyorder=p)




# ------------------------------------------------------------
# Sliding-window modulus
# ------------------------------------------------------------
def compute_sliding_window_modulus(fail_df, CSA_true,
                                   window_size_strain=0.01):
    """
    

    window_size_strain : strain window width (default 0.01 = ±0.005)
    
    Returns:
        modulus,
        strain_at_mod,
        stress_at_mod,
        global_index
    """

    import numpy as np

    # Extract arrays
    force = fail_df["Force_N"].to_numpy(float)
    disp  = fail_df["Displacement_mm"].to_numpy(float)
    size  = float(fail_df["Size_mm"].iloc[0])

    # Baseline correction:
    load_corr = force - force[0]
    disp_corr = disp - disp[0]

    strain_raw = disp_corr / size
    stress_raw = load_corr / CSA_true if CSA_true != 0 \
                 else np.full_like(load_corr, np.nan)

    mask = np.isfinite(strain_raw) & np.isfinite(stress_raw)
    strain = strain_raw[mask]
    stress = stress_raw[mask]
    idx_all = fail_df.index.to_numpy()[mask]

    if len(strain) < 6:
        return np.nan, np.nan, np.nan, None

    # Pre-peak region
    peak_idx = int(np.nanargmax(stress))
    if peak_idx <= 2:
        return np.nan, np.nan, np.nan, None

    strain = strain[:peak_idx]
    stress = stress[:peak_idx]
    idx_all = idx_all[:peak_idx]

    if len(strain) < 6:
        return np.nan, np.nan, np.nan, None

    # --------------------------------------------------------
    # Peak-stress window filter (e.g. 5%–60% of peak stress)
    # --------------------------------------------------------
    peak_stress = float(np.nanmax(stress))
    low_frac  = 0.05     # lower 5%
    high_frac = 0.60     # upper 60%

    low_s  = low_frac  * peak_stress
    high_s = high_frac * peak_stress

    mask_zone = (stress >= low_s) & (stress <= high_s)

    if np.sum(mask_zone) < 6:
        # Not enough points – sliding modulus cannot be computed
        return np.nan, np.nan, np.nan, None

    strain = strain[mask_zone]
    stress = stress[mask_zone]
    idx_all = idx_all[mask_zone]


    # Sliding-linear behavior
    n = len(strain)
    slopes = np.full(n, np.nan)
    half_win = window_size_strain / 2.0

    for i in range(n):
        x0 = strain[i]
        mask_win = (strain >= x0 - half_win) & (strain <= x0 + half_win)
        xs = strain[mask_win]
        ys = stress[mask_win]

        if len(xs) < 3:
            continue

        try:
            m, c = np.polyfit(xs, ys, 1)
            slopes[i] = m
        except Exception:
            continue

    if not np.isfinite(slopes).any():
        return np.nan, np.nan, np.nan, None

    idx_local = int(np.nanargmax(slopes))
    global_idx = int(idx_all[idx_local])

    modulus = float(slopes[idx_local])
    strain_at = float(strain[idx_local])
    stress_at = float(stress[idx_local])

    return modulus, strain_at, stress_at, global_idx



# ---------------------------------------------------------------------
# Analysis of a single file
# ---------------------------------------------------------------------

def analyze_file(csv_path, metadata):
    print(f"Processing {csv_path} ...")

    date_id, sample_id, rep = parse_sample_from_filename(csv_path)
    if not date_id:
        print("  ! Could not parse filename.")
        return None

    meta = metadata[
        (metadata["Date_ID"].astype(str) == str(date_id)) &
        (metadata["Sample_ID"].astype(str) == str(sample_id)) &
        (metadata["Replicate"].astype(str) == str(rep))
    ]
    if meta.empty:
        print("  ! Metadata not found.")
        return None

    meta = meta.iloc[0]

    date_val = meta["Date"]
    sex = meta["Sex"]
    age = meta["Age"]
    genotype = meta["Genotype"]

    avg_diam = float(meta["Average_diameter (um)"])
    CSA_raw  = float(meta["C.s.a (um squared)"])
    CSA_true = float(meta["C.s.a (mm squared)"])

    df = pd.read_csv(csv_path)
    req = ["SetName","Cycle","Time_S","Size_mm","Displacement_mm","Force_N"]
    for c in req:
        if c not in df.columns:
            raise ValueError(f"Column {c} missing.")

    # PRECONDITIONING ----------------------------------------------------
    pre = df[df["SetName"].str.contains("5x pre-conditioning", case=False, na=False)].copy()
    if pre.empty:
        print("  ! No preconditioning.")
        return None

    sample_length = float(pre["Size_mm"].iloc[0])
    min_force = float(pre["Force_N"].min())

    pre["Load_corr"] = pre["Force_N"] - min_force
    pre["Disp_corr"] = pre["Displacement_mm"] - pre["Displacement_mm"].iloc[0]

    pre["Load_smooth"] = apply_savgol_safe(pre["Load_corr"], PRECON_SAVGOL_WINDOW, PRECON_SAVGOL_POLY)
    pre["Disp_smooth"] = apply_savgol_safe(pre["Disp_corr"], PRECON_SAVGOL_WINDOW, PRECON_SAVGOL_POLY)

    if CSA_true != 0:
        pre["Stress_smooth"] = pre["Load_smooth"] / CSA_true
    else:
        pre["Stress_smooth"] = np.nan

    c1 = pre["Cycle"].astype(str).str.startswith("1-")


    def trapz_inc(x, y):
        x = np.asarray(x, float)
        y = np.asarray(y, float)
        dx = x[1:] - x[:-1]
        inc = 0.5 * (y[1:] + y[:-1]) * dx
        return np.concatenate([[0], np.cumsum(inc)])


    pre["Area_smooth"] = trapz_inc(pre["Disp_smooth"], pre["Load_smooth"])

    # --- Classical single-cycle hysteresis for Cycle 1 (Stretch vs Recover) ---

    # Identify stretch and recover phases in Cycle 1
    c1_stretch = pre["Cycle"].astype(str).str.contains("1-Stretch", case=False)
    c1_recover = pre["Cycle"].astype(str).str.contains("1-Recover", case=False)

    if c1_stretch.any() and c1_recover.any():
        # Stretch phase (loading)
        xs_load = pre.loc[c1_stretch, "Disp_smooth"].to_numpy(float)
        ys_load = pre.loc[c1_stretch, "Load_smooth"].to_numpy(float)
        xs_load = xs_load - xs_load[0]   # zero displacement baseline

        # Recover phase (unloading)
        xs_unload = pre.loc[c1_recover, "Disp_smooth"].to_numpy(float)
        ys_unload = pre.loc[c1_recover, "Load_smooth"].to_numpy(float)
        xs_unload = xs_unload - xs_load[0]  # use same reference for unloading

        # Integrate areas
        A_load_classical = float(np.trapz(ys_load, xs_load))
        A_unload_classical = float(np.trapz(ys_unload, xs_unload))

        # Classical hysteresis (%)
        hyst_energy_classical = A_load_classical - abs(A_unload_classical)
        hyst_pct_classical = hyst_energy_classical / A_load_classical * 100 if A_load_classical != 0 else np.nan
    else:
        hyst_energy_classical = np.nan
        hyst_pct_classical = np.nan




    # HOLD 60s --------------------------------------------------------------
    hold = df[df["Cycle"].astype(str).str.contains("1-Hold", case=False, na=False)].copy()

    stress_relax_60 = np.nan
    stress_rate_60  = np.nan

    if not hold.empty and CSA_true != 0:
        hold["Stress_raw"] = hold["Force_N"] / CSA_true
        hold["Stress_smooth"] = apply_savgol_safe(hold["Stress_raw"], HOLD_SAVGOL_WINDOW, HOLD_SAVGOL_POLY)

        times = hold["Time_S"].to_numpy(float)
        stress_sm = hold["Stress_smooth"].to_numpy(float)

        if len(times) > 1:
            t0 = times[0]
            t60 = t0 + 60.0
            if times[-1] >= t60:
                s0 = stress_sm[0]
                s60 = float(np.interp(t60, times, stress_sm))

                if s0 != 0:
                    stress_relax_60 = (s0 - s60)/s0 * 100
                stress_rate_60 = (s60 - s0)/60.0

    # FAILURE REGION  ----------------------------------------
    fail = df[
        df["SetName"].astype(str).str.lower().str.contains("fail")
    ].copy()

    if fail.empty:
        print("  ! No failure segment found (SetName did not contain 'fail').")
        return None


    fail["Load_corr"] = fail["Force_N"] - fail["Force_N"].iloc[0]
    fail["Disp_corr"] = fail["Displacement_mm"] - fail["Displacement_mm"].iloc[0]
    fail["Strain"] = fail["Disp_corr"] / sample_length

    if CSA_true != 0:
        fail["Stress_MPa"] = fail["Load_corr"] / CSA_true
    else:
        fail["Stress_MPa"] = np.nan

    # create plot directory & save PNG
    plot_dir = "plots"
    os.makedirs(plot_dir, exist_ok=True)

    plot_path = os.path.join(
        plot_dir,
        f"{os.path.basename(csv_path)}_force_time.png"
    )
    save_force_time_plot(df, plot_path)
    

    # CORRECTED -------------------------------------------------------------
    fidx_corr = fail["Load_corr"].idxmax()
    failure_force_corr = float(fail.loc[fidx_corr, "Load_corr"])
    failure_ext_corr = float(fail.loc[fidx_corr, "Disp_corr"])
    failure_strain_corr = float(fail.loc[fidx_corr, "Strain"])
    failure_stress_corr = float(fail.loc[fidx_corr, "Load_corr"] / CSA_true) if CSA_true != 0 else np.nan


    # ---------- Sliding-window modulus ----------
    max_mod_poly = np.nan
    stress_at_mod_poly = np.nan
    strain_at_mod_poly = np.nan

    time_at_mod_poly = np.nan
    force_at_mod_poly = np.nan
    disp_at_mod_poly = np.nan

    # detection of the failure region
    fail_poly = df[df["SetName"].astype(str).str.lower().str.contains("fail")].copy()

    if not fail_poly.empty and CSA_true != 0:

        (m_lin, strain_lin, stress_lin, global_idx) = compute_sliding_window_modulus(
    fail_poly,
    CSA_true
)


        if global_idx is not None:
            max_mod_poly       = m_lin
            strain_at_mod_poly = strain_lin
            stress_at_mod_poly = stress_lin

            force_at_mod_poly = float(fail_poly["Force_N"].loc[global_idx])
            time_at_mod_poly  = float(fail_poly["Time_S"].loc[global_idx])
            disp_at_mod_poly  = float(fail_poly["Displacement_mm"].loc[global_idx])


    # Produce Force–Extension plot (with X marking data from which modulus is derived)
    force_extension_plot_path = os.path.join(
        plot_dir,
        f"{os.path.basename(csv_path)}_force_extension.png"
    )

    save_force_extension_plot(
        fail_poly,
        force_extension_plot_path,
        disp_mod=disp_at_mod_poly if np.isfinite(disp_at_mod_poly) else None,
        force_mod=force_at_mod_poly if np.isfinite(force_at_mod_poly) else None
    )



    # ---------------------------------------------------------------------
    # Assemble results
    # ---------------------------------------------------------------------

    result = {
        "File name": os.path.basename(csv_path),
        "Date": date_val,
        "Sample ID": sample_id,
        "Replicate number": int(rep),
        "Sex": sex,
        "Age": age,
        "Genotype": genotype,


        "Stress-relaxation 60s": stress_relax_60,
        "Rate of change of stress 60s": stress_rate_60,

        "Hysteresis energy (cycle 1 classical)": hyst_energy_classical,
        "Hysteresis % (cycle 1 classical)": hyst_pct_classical,

        "Average diameter (um)": avg_diam,
        "C.s.a (um squared)": CSA_raw,
        "C.s.a (mm squared)": CSA_true,

        "Failure force (N) – corrected": failure_force_corr,
        "Failure stress (MPa) – corrected": failure_stress_corr,
        "Failure strain (%) – corrected": failure_strain_corr * 100,
        "Failure extension (mm) – corrected": failure_ext_corr,

        # Sliding modulus
        "Max modulus (sliding window)": max_mod_poly,
        "Stress at max modulus (sliding window)": stress_at_mod_poly,
        "Strain at max modulus (sliding window)": strain_at_mod_poly,

        # file path for plot
        "Force-Time plot": plot_path,
        "Force-Extension plot": force_extension_plot_path,

    }

    # Convert numpy scalars
    for k, v in result.items():
        if isinstance(v, np.generic):
            result[k] = v.item()

    return result

# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main():
    metadata = load_metadata("tendon_data_v4.csv")
    files = find_data_files(".")
    print(f"Found {len(files)} files.")

    results = []
    for path in files:
        try:
            r = analyze_file(path, metadata)
            if r is not None:
                results.append(r)
        except Exception as e:
            print(f"  !! Error processing {path}: {e}")

    if not results:
        print("No valid samples.")
        return

    df = pd.DataFrame(results)

    # ---------- FINAL CLEAN ----------
    for col in df.columns:

        if col in ["Age", "Sex", "Genotype", "Sample ID", "Replicate number"]:
            continue

        df[col] = (
            df[col]
            .astype(str)
            .str.strip()
            .str.replace("−", "-", regex=False)
            .str.replace("–", "-", regex=False)
            .str.replace("nan", "", regex=False)
            .replace("", np.nan)
        )
        df[col] = pd.to_numeric(df[col], errors="ignore")

    print("\nCOLUMN TYPES:")
    print(df.dtypes)

    # Save CSV (cannot contain images)
    df.to_csv(
        "results_summary.csv",
        index=False,
        float_format="%.10g"
    )
    print("✔ results_summary.csv written.")

    # Save XLSX first, without images
    xlsx_path = "results_summary.xlsx"
    df.to_excel(xlsx_path, index=False)
    print("✔ results_summary.xlsx written (pre-image stage).")

    # Embed images into Excel
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = list(df.columns)
    plot_col_idx = headers.index("Force-Time plot") + 1  # 1-based Excel column index
    force_ext_col_idx = headers.index("Force-Extension plot") + 1

    for row_idx in range(2, len(df) + 2):  # Excel data rows
        plot_path = ws.cell(row=row_idx, column=plot_col_idx).value
        if plot_path and os.path.exists(plot_path):

            # Insert image
            img = XLImage(plot_path)
            img.width = 300
            img.height = 225

            cell = f"{get_column_letter(plot_col_idx)}{row_idx}"
            ws.add_image(img, cell)

            # --- adjust row height so image fits completely ---
            ws.row_dimensions[row_idx].height = 170  # approx. height for a 225px image

    # — Insert Force–Extension images
    for row_idx in range(2, len(df) + 2):
        plot_path = ws.cell(row=row_idx, column=force_ext_col_idx).value
        if plot_path and os.path.exists(plot_path):

            img = XLImage(plot_path)
            img.width = 300
            img.height = 225

            cell = f"{get_column_letter(force_ext_col_idx)}{row_idx}"
            ws.add_image(img, cell)

            ws.row_dimensions[row_idx].height = 170


    wb.save(xlsx_path)
    print("✔ Images embedded into results_summary.xlsx.")


if __name__ == "__main__":
    main()
